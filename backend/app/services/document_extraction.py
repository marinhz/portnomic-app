"""Extract text from uploaded documents (PDF, XLSX, JPG) for AI parsing."""

from __future__ import annotations

import io
import logging
from pathlib import Path

logger = logging.getLogger("shipflow.document_extraction")

ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".jpg", ".jpeg"}
ALLOWED_CONTENT_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "image/jpeg",
}
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB


def _get_extension(filename: str) -> str:
    return Path(filename).suffix.lower()


def validate_file(filename: str, content_type: str | None, size: int) -> None:
    """Validate file type and size. Raises ValueError if invalid."""
    ext = _get_extension(filename)
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type. Allowed: PDF, XLSX, JPG. Got: {ext or 'unknown'}"
        )
    if size > MAX_FILE_SIZE_BYTES:
        raise ValueError(
            f"File too large. Maximum {MAX_FILE_SIZE_BYTES // (1024*1024)} MB."
        )
    if content_type and content_type.lower() not in ALLOWED_CONTENT_TYPES:
        # Lenient: content-type can be wrong (e.g. multipart); extension is primary
        pass


def extract_text_from_document(data: bytes, filename: str) -> str:
    """Extract text from a document for LLM parsing.

    Supports PDF, XLSX, JPG. For JPG (scanned documents), requires pytesseract.
    Raises ValueError for unsupported types or extraction failure.
    """
    ext = _get_extension(filename)
    stream = io.BytesIO(data)

    if ext == ".pdf":
        return _extract_pdf(stream)
    if ext in (".xlsx", ".xls"):
        return _extract_xlsx(stream)
    if ext in (".jpg", ".jpeg"):
        return _extract_jpg(stream)

    raise ValueError(f"Unsupported file type: {ext}")


def _extract_pdf(stream: io.BytesIO) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ValueError(
            "PDF extraction requires pypdf. Install with: pip install pypdf"
        ) from None

    reader = PdfReader(stream)
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    if not parts:
        raise ValueError("PDF contains no extractable text (may be image-only)")
    return "\n\n".join(parts)


def _extract_xlsx(stream: io.BytesIO) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError(
            "XLSX extraction requires openpyxl. Install with: pip install openpyxl"
        ) from None

    wb = load_workbook(stream, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                rows.append("\t".join(vals))
        if rows:
            parts.append("\n".join(rows))
    wb.close()
    if not parts:
        raise ValueError("XLSX contains no readable data")
    return "\n\n--- Sheet ---\n\n".join(parts)


def _extract_jpg(stream: io.BytesIO) -> str:
    """Extract text from JPG using OCR (pytesseract)."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        raise ValueError(
            "JPG/OCR extraction requires pytesseract and Pillow. "
            "Install: pip install pytesseract pillow. "
            "Also install Tesseract OCR: https://github.com/tesseract-ocr/tesseract"
        ) from None

    try:
        img = Image.open(stream)
        img.load()
    except Exception as e:
        raise ValueError(f"Failed to load image: {e}") from e

    try:
        text = pytesseract.image_to_string(img)
    except Exception as e:
        raise ValueError(
            f"OCR failed. Is Tesseract installed? Error: {e}"
        ) from e

    if not text or not text.strip():
        raise ValueError("Image contains no recognizable text")
    return text.strip()


__all__ = [
    "extract_text_from_document",
    "validate_file",
    "ALLOWED_EXTENSIONS",
    "ALLOWED_CONTENT_TYPES",
]
